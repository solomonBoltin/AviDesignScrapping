import os
from asyncio import sleep

from openpyxl import load_workbook, Workbook

from pages.WallpaperPage import WallpaperPage

ID = "A"
TITLE = "B"
PRICE = "C"
DESCRIPTION = "D"
CATEGORIES = "E"
TAGS = "F"
IMAGES = "G"
LINK = "H"


class WallpapersDatabase:
    file_path = "./data/wallpapers.xlsx"

    def __init__(self):
        if os.path.isfile(self.file_path):
            os.remove(self.file_path)
        self.workbook = self.verify_file()

    def add_wallpaper(self, wallpaper_page: WallpaperPage):
        wallpaper_id = wallpaper_page.product_id
        wallpaper_title = wallpaper_page.title
        wallpaper_price = wallpaper_page.price
        wallpaper_details = wallpaper_page.details
        wallpaper_categories = wallpaper_page.categories
        wallpaper_tags = wallpaper_page.hashtags
        wallpaper_images = wallpaper_page.images
        wallpaper_page_url = wallpaper_page.page_url

        product_row = self.products_len + 1
        self.workbook.active[f"{ID}{product_row}"] = wallpaper_id
        self.workbook.active[f"{TITLE}{product_row}"] = wallpaper_title
        self.workbook.active[f"{PRICE}{product_row}"] = wallpaper_price
        self.workbook.active[f"{DESCRIPTION}{product_row}"] = wallpaper_details
        self.workbook.active[f"{CATEGORIES}{product_row}"] = wallpaper_categories
        self.workbook.active[f"{TAGS}{product_row}"] = wallpaper_tags
        self.workbook.active[f"{IMAGES}{product_row}"] = wallpaper_images
        self.workbook.active[f"{LINK}{product_row}"] = wallpaper_page_url

        self.workbook.save(self.file_path)

    def verify_file(self):
        try:
            return load_workbook(filename=self.file_path)
        except FileNotFoundError:
            return self.create_xlsx_file()

    def create_xlsx_file(self):
        workbook = Workbook()
        workbook.active[f"{ID}{1}"] = "ID"
        workbook.active[f"{TITLE}{1}"] = "TITLE"
        workbook.active[f"{PRICE}{1}"] = "PRICE"
        workbook.active[f"{DESCRIPTION}{1}"] = "DESCRIPTION"
        workbook.active[f"{CATEGORIES}{1}"] = "CATEGORIES"
        workbook.active[f"{TAGS}{1}"] = "TAGS"
        workbook.active[f"{IMAGES}{1}"] = "IMAGES"
        workbook.active[f"{LINK}{1}"] = "LINK"

        workbook.save(self.file_path)

        return workbook

    @property
    def products_len(self):
        return self.workbook.worksheets[0].max_row
